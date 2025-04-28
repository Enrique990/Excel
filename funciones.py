
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import datetime
import openpyxl.utils

def cargar_notas():
    try:
        wb = openpyxl.load_workbook('notas_estudiantes.xlsx')
    except FileNotFoundError:
        print("No se encontro el archivo de notas")
        return None
    
    ws = wb.active
    ws.title = "notas"
    return  wb, ws

def automatizacion_notas():
    wb, ws = cargar_notas()
    if not wb:
        return
    
    wb = generar_reporte()
    wb.save('notas_estudiantes.xlsx')
    print("\nProceso completado. Resultados mostrados arriba y archivo guardado.")

def calcular_estadisticas(ws):
    """Calcula estadísticas desde los datos y las imprime en consola"""
    notas = [cell.value for row in ws.iter_rows(min_row=2, min_col=2, max_col=2) 
             for cell in row if isinstance(cell.value, (int, float))]
    
    if not notas:
        print("No se encontraron notas válidas")
        return None
    
    total = len(notas)
    aprobados = sum(1 for n in notas if n >= 70)
    reprobados = total - aprobados
    reprobados_60_69 = sum(1 for n in notas if 60 <= n < 70)
    media = sum(notas)/total
    varianza = sum((n-media)**2 for n in notas)/total
    desviacion = varianza**0.5
    
    # Imprimir resultados en consola
    print("\n=== RESULTADOS CALCULADOS EN PYTHON ===")
    print(f"1. Número total de estudiantes: {total}")
    print(f"2. Estudiantes aprobados (nota >= 70): {aprobados} ({aprobados/total*100:.2f}%)")
    print(f"3. Estudiantes reprobados (nota < 70): {reprobados} ({reprobados/total*100:.2f}%)")
    print(f"4. Reprobados con notas entre 60-69: {reprobados_60_69} ({reprobados_60_69/total*100:.2f}%)")
    print(f"5. Media de las notas: {media:.2f}")
    print(f"6. Desviación estándar de las notas: {desviacion:.2f}")
    
    return {
        'total': total,
        'aprobados': aprobados,
        'porc_aprobados': aprobados/total,
        'reprobados': reprobados,
        'porc_reprobados': reprobados/total,
        'reprobados_60_69': reprobados_60_69,
        'porc_reprobados_60_69': reprobados_60_69/total,
        'media': media,
        'desviacion': desviacion
    }

def generar_reporte():
    wb, ws = cargar_notas()
    if not wb:
        return
    
    calcular_estadisticas(ws)
    
    #crear una hoja nueva para el reporte
    fecha_actual = datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")
    ws_reporte = wb.create_sheet(f"Reporte {fecha_actual}")

    #agregar encabezados
    ws_reporte.cell(row=1, column=1, value='Estadisticas').font = Font(bold=True)
    ws_reporte.cell(row=1, column=2, value='Numeros').font = Font(bold=True)
    ws_reporte.cell(row=1, column=3, value='Porcentajes').font = Font(bold=True)
    ws_reporte.cell(row=2, column=1, value='Numero total de estudiantes')
    ws_reporte.cell(row=3, column=1, value='Numero de estudiantes aprobados (nota >= 70)')
    ws_reporte.cell(row=4, column=1, value='Porcentaje de estudiantes aprobados (nota >= 70)')
    ws_reporte.cell(row=5, column=1, value='Numero de estudiantes reprobados (nota < 70)')
    ws_reporte.cell(row=6, column=1, value='Porcentaje de estudiantes reprobados (nota < 70)')
    ws_reporte.cell(row=7, column=1, value='Numero de estudiantes reprobados con notas entre 60 y 69')
    ws_reporte.cell(row=8, column=1, value='Porcentaje de estudiantes reprobados con notas entre 60 y 69')
    ws_reporte.cell(row=9, column=1, value='Media de las notas')
    ws_reporte.cell(row=10, column=1, value='Desviación estandar de las notas')

    #calcular total de estudiantes
    ws_reporte.cell(row=2, column=2, value=f"=COUNTA(notas!A2:A{ws.max_row})")

    #calcular numero de estudiantes aprobados (nota >= 70)
    ws_reporte.cell(row=3, column=2, value=f'=COUNTIF(notas!B2:B{ws.max_row},">=70")')

    #calcular porcentaje de estudiantes aprobados (nota >= 70)
    ws_reporte.cell(row=4, column=3, value=f"=B3/B2")
    ws_reporte["C4"].number_format = '0.00%'

    #calcular numero de estudiantes reprobados (nota < 70)
    ws_reporte.cell(row=5, column=2, value=f'=COUNTIF(notas!B2:B{ws.max_row},"<70")')

    #calcular porcentaje de estudiantes reprobados (nota < 70)
    ws_reporte.cell(row=6, column=2, value=f"=B5")  # opcional, repetir valor absoluto
    ws_reporte.cell(row=6, column=3, value=f"=B5/B2")
    ws_reporte["C6"].number_format = '0.00%'

    #calcular numero de estudiantes reprobados con notas entre 60 y 69
    ws_reporte.cell(row=7, column=2, value=f'=COUNTIFS(notas!B2:B{ws.max_row},">=60",notas!B2:B{ws.max_row},"<70")')

    #calcular porcentaje de estudiantes reprobados con notas entre 60 y 69
    ws_reporte.cell(row=8, column=2, value=f"=B7")  # opcional, solo para mantener consistencia
    ws_reporte.cell(row=8, column=3, value=f"=B7/B2")
    ws_reporte["C8"].number_format = '0.00%'

    #calcular media de las notas
    ws_reporte.cell(row=9, column=2, value=f"=AVERAGE(notas!B2:B{ws.max_row})")
    ws_reporte["B9"].number_format = '0.00'

    #calcular desviación estandar de las notas
    ws_reporte.cell(row=10, column=2, value=f"=STDEVP(notas!B2:B{ws.max_row})")
    ws_reporte["B10"].number_format = '0.00'

    ws_reporte.column_dimensions[openpyxl.utils.get_column_letter(1)].width = 50
    ws_reporte.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 11

    return wb